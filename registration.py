# import openpyxl and tkinter modules
import openpyxl
from tkinter import *
from tkinter import filedialog

# Create global variables for the workbook and sheet
# None is there because to ensure that they have a default value and error will not come 
# up when we use these variables directly in the workbook and sheet functions. 
wb = None
sheet = None
		
# opening the existing excel file
wb = openpyxl.load_workbook("C:\\Users\\LENOVO\\Desktop\\vs_coding\\Python\\excel.xlsx")
 
# create the sheet object and active it
# in this line this ensures that the active sheet in the Excel file is selected and can be manipulated by the code.
sheet = wb.active


# Define a function to format the Excel sheet
def excel():
    if sheet is None:
        print("Sheet is not initialized. Please open an Excel file first.")
        return

    # resize the width of columns in excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

# Define functions to set focus on text entry boxes when the Enter key is pressed

# event argument is set to a string containing the name of the text entry box 
# and is the object in Tkinter module
	
# Function to set focus (cursor)
# The function uses the focus_set() method to set the focus to the course_field text entry box.
# This method sets the cursor to the text entry box, allowing the user to start typing in it.
def focus1(event):
	# set focus on the course_field box
	course_field.focus_set()

# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	sem_field.focus_set()

# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	form_no_field.focus_set()

# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	contact_no_field.focus_set()

# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	email_id_field.focus_set()

# Function to set focus
def focus6(event):
	# set focus on the address_field box
	address_field.focus_set()


# Function for clearing the contents of text entry boxes in GUI window
# It is called at the end of the inset function when data is saved
# It happens when we click the submit button on the GUI window.
def clear():
	# clear the content of text entry box for the next entry.
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)
# It uses the delete method of the Entry widget to delete the text from the text entry boxes.
# The delete method takes two arguments: the starting index and the ending index of the text to be deleted.
# In this case, we are deleting the entire text by passing 0 as the starting index and END as the ending index.

# Function to take data from GUI window and write to an excel files
def insert():

	# This if condition will check all entry boxes are empty or not
	# if it is empty then print "empty input"
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")
		
	else:
		# assigning the max row and max column value upto which data is written in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text as string which we write into excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		# This line selects the cell in the Excel sheet where the name value should be written. 
		# The current_row variable keeps track of the current row number in the Excel sheet.
		# By using current_row + 1, we're selecting the next available row in the sheet. 
		# The column=1 part specifies that we want to write to the first column in the sheet (since column numbers start from 0).
		
        # The current_row variable is incremented after writing the data to the Excel file 
		# to ensure that the next time the "Submit" button is clicked, the data will be written to the next row in the Excel sheet.
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=7).value = address_field.get()

		# save the file
		wb.save("C:\\Users\\LENOVO\\OneDrive\\Desktop\\excel.xlsx")

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()

# Driver code
		
# This if condition is used to determine the execution context of the script.
		
    # If the script is run directly, the __name__ variable will be set to "__main__", and the code within the if statement will be executed. 
    # If the script is imported as a module in another script, the __name__ variable will be set to the name of the script, 
	# and the code within the if statement will not be executed.
            
    # if we remove the line if __name__ == "__main__":, the code within the if statement will always be executed, 
    # regardless of whether the script is run directly or imported as a module. 
    # This may not be desirable, especially if the script contains code that is intended to be run only when the script is run directly.
    # so to ensure that the script behaves as intended in different execution contexts we will write it.

if __name__ == "__main__":
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='light green')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("500x300")
    
	excel()

	# create a Form label
	heading = Label(root, text="Form", bg="light green")

	# create a Name label
	name = Label(root, text="Name", bg="light green")

	# create a Course label
	course = Label(root, text="Course", bg="light green")

	# create a Semester label
	sem = Label(root, text="Semester", bg="light green")

	# create a Form No. label
	form_no = Label(root, text="Form No.", bg="light green")

	# create a Contact No. label
	contact_no = Label(root, text="Contact No.", bg="light green")

	# create a Email id label
	email_id = Label(root, text="Email id", bg="light green")

	# create a address label
	address = Label(root, text="Address", bg="light green")

	# grid method is used for placing the widgets at respective positions in table like structure .
	heading.grid(row=0, column=1)
	# This line places the label widget in the GUI window at row 1, column 0.
	name.grid(row=1, column=0)
	course.grid(row=2, column=0)
	sem.grid(row=3, column=0)
	form_no.grid(row=4, column=0)
	contact_no.grid(row=5, column=0)
	email_id.grid(row=6, column=0)
	address.grid(row=7, column=0)

	# create a text entry box for typing the information
	name_field = Entry(root)
	course_field = Entry(root)
	sem_field = Entry(root)
	form_no_field = Entry(root)
	contact_no_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)

	# bind method of widget is used for the binding the function with the events

	# whenever the enter key is pressed then call the focus1 function
	name_field.bind("<Return>", focus1)
	
	# whenever the enter key is pressed then call the focus2 function
	course_field.bind("<Return>", focus2)

	# whenever the enter key is pressed then call the focus3 function
	sem_field.bind("<Return>", focus3)

	# whenever the enter key is pressed then call the focus4 function
	form_no_field.bind("<Return>", focus4)

	# whenever the enter key is pressed then call the focus5 function
	contact_no_field.bind("<Return>", focus5)

	# whenever the enter key is pressed then call the focus6 function
	email_id_field.bind("<Return>", focus6)


	# grid method is used for placing the widgets at respective positions in table like structure .
	# ipadx is an option in the grid() method that adds internal padding to the x-axis (width) of the widget. 
	# In this case, it adds 100 pixels of internal padding to the width of the text entry boxes. 
	# This means that the text entry boxes will be 100 pixels wider than their default size, providing more space for the user to type in.
	name_field.grid(row=1, column=1, ipadx="100")
	course_field.grid(row=2, column=1, ipadx="100")
	sem_field.grid(row=3, column=1, ipadx="100")
	form_no_field.grid(row=4, column=1, ipadx="100")
	contact_no_field.grid(row=5, column=1, ipadx="100")
	email_id_field.grid(row=6, column=1, ipadx="100")
	address_field.grid(row=7, column=1, ipadx="100")

	# call excel function
	excel()

	# create a Submit Button and place into the root window
	submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert)
	submit.grid(row=8, column=1)

	# start the GUI
	root.mainloop()